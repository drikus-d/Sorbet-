# excel_format.py

import pandas as pd
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from typing import List, Tuple, Dict
from datetime import datetime, timedelta
import os
import config
from ops_kpi import combine_ops_kpi_dataframes
from decimal import Decimal

def preprocess_data_for_excel(df):
    """
    Preprocess DataFrame to convert percentage strings back to raw decimals for Excel formatting
    """
    df_copy = df.copy()
    
    # Define columns that contain percentage data
    percentage_columns = ['Green', 'Blue', 'Silver', 'Gold', 'Total Loyalty', 'Non-loyalty', 'Total']
    
    for col in percentage_columns:
        if col in df_copy.columns:
            # Convert percentage strings back to raw decimals
            df_copy[col] = df_copy[col].apply(lambda x: convert_percentage_to_decimal(x))
    
    return df_copy

def convert_percentage_to_decimal(value):
    """
    Convert percentage string (like "1.6%") back to decimal (like 0.016)
    """
    if isinstance(value, str) and '%' in value:
        try:
            # Remove % and convert to float, then divide by 100
            return float(value.replace('%', '')) / 100
        except (ValueError, AttributeError):
            return value
    return value

def load_template(template_path: str) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"The template file was not found at: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active  # Modify if you want a specific sheet
    return wb, ws

def write_metadata(ws, start_date: str, end_date: str, execution_time: str, cell: str = "A2"):
    metadata_text = f"For {start_date} to {end_date}    Report Execution Time: {execution_time}"
    ws[cell] = metadata_text
    ws[cell].font = Font(bold=True)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def create_black_border():
    black_side = Side(border_style="thin", color="000000")
    return Border(
        left=black_side,
        right=black_side,
        top=black_side,
        bottom=black_side
    )

def write_headers(ws, row: int, columns: List[str], exception_subheaders: List[str], current_subheader: str):
    # Check if current subheader is in the list of exceptions
    is_exception = current_subheader in exception_subheaders

    # Apply black fill to both "Metric" and "Type" columns for "Loyalty KPI"
    is_loyalty_kpi = current_subheader == "Loyalty KPI"

    # Write column headers starting from the first column
    for col_num, column_title in enumerate(columns, start=1):
        # Skip formatting for separator columns
        if isinstance(column_title, str) and column_title.startswith('separator_'):
            header_cell = ws.cell(row=row, column=col_num+1, value="")
            header_cell.border = None  # Remove borders
            header_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_cell.font = Font(bold=True)
            continue
        
        # Add check for top left column to make blank
        if column_title=='Service Items Comparison' or column_title=='Retail Items Comparison':
            header_cell = ws.cell(row=row, column=col_num+1, value='')
        else:
            header_cell = ws.cell(row=row, column=col_num+1, value=column_title)
        
        # # Apply black fill for "Metric" and "Type" columns if this is "Loyalty KPI"
        # if (col_num == 1 or (col_num == 2 and is_loyalty_kpi)) and not is_exception:
        #     header_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black fill
        # else:
        #     header_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")  # Standard fill
        if 'Weekly Performance' in str(ws.cell(row=row-1, column=2).value):
            if col_num == 1:
                header_cell.font = Font(bold=True,color="FFFFFF")
                header_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            else:
                header_cell.font = Font(bold=True,color="FFFFFF")
                header_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        elif 'Total Service Items' in str(ws.cell(row=row-1, column=2).value) or 'Total Retail Items' in str(ws.cell(row=row-1, column=2).value):
            if column_title=='Top 20 Previous Units Over Total %':
                ws.cell(row=row-1, column=2).value = ''
            if column_title=='Service Items Comparison' or column_title=='Retail Items Comparison':
                header_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            else:
                header_cell.font = Font(bold=True,color="FFFFFF")
                header_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")                
        else:
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_cell.border = create_black_border()  # Apply black border to header cells

def write_data(ws, row: int, df: pd.DataFrame, percentage_columns: List[str], two_decimal_columns: List[str],
               merge_columns: List[str] = None, per_row_formatting: Dict[str, Dict[str, Dict[str, str]]] = None,
               exception_metric_types: List[Tuple[str, str]] = None, subheader: str = None) -> int:
    # Preprocess data for Excel - convert percentage strings back to decimals
    df = preprocess_data_for_excel(df)
    
    # If merge_columns is provided, keep track of merge ranges
    merge_ranges = {col: {'start_row': row, 'value': None} for col in merge_columns} if merge_columns else {}

    for idx, data_row in df.iterrows():
        # Get the metric name and clean it by removing indentation and extra spaces
        metric_name = str(data_row.get('Metric_Type' if 'Metric_Type' in df.columns else 'Metric' if 'Metric' in df.columns else 'act_bud')).strip()
        clean_metric_name = ' '.join(metric_name.split())  # This removes extra spaces and indentation
        

        # Check if the row is a % row
        is_percent_row = 0

        if "%" in str(data_row[0]).strip():
            is_percent_row = 1
        
        # # Check if this is the "Birthday discount cost % of sales" row with type "Current"
        # is_birthday_cost_current = (
        #     data_row.get('Metric') == 'Birthday discount cost % of sales' and 
        #     data_row.get('Type') == 'Current'
        # )

        for col_num, (column_title, value) in enumerate(data_row.items(), 1):
            cell = ws.cell(row=row, column=col_num+1)
            cell.value = value
            cell.border = create_black_border()  # Apply black border to data cells

            # # Bold formatting for "Metric" column in "Loyalty KPI" section
            # if column_title == "Metric":
            #     cell.font = Font(bold=True)

            # if merge_columns and column_title in merge_columns:
            #     # Handle merge tracking but don't apply formatting yet
            #     continue

            # Apply metric-based formatting for data columns
            if clean_metric_name in config.OPS_METRIC_FORMAT_MAPPING and column_title not in ['Metric_Type', 'Metric', 'Type']:
                cell.number_format = config.OPS_METRIC_FORMAT_MAPPING[clean_metric_name]
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                
                # Special handling for percentage metrics - convert values to decimal format for Excel
                if '0.0%' in config.OPS_METRIC_FORMAT_MAPPING[clean_metric_name]:
                    if pd.notna(value) and value != '':
                        try:
                            numeric_value = float(value)
                            # Convert from percentage (1.27) to decimal (0.0127) for Excel's 0.0% format
                            cell.value = numeric_value / 100
                            
                            # Apply background color based on original value
                            if numeric_value >= 0:
                                cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")  # Light green
                            else:
                                cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
                        except (ValueError, TypeError):
                            cell.value = value
            # Special handling for Basket Size percentage rows
            elif subheader in ['Basket Size Total', 'Basket Size by Region'] and '%' in clean_metric_name and column_title not in ['Metric_Type', 'Metric', 'Type', 'act_bud']:
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                
                if pd.notna(value) and value != '':
                    try:
                        numeric_value = float(value)
                        # Convert from percentage (1.27) to decimal (0.0127) for Excel's 0.0% format
                        cell.value = numeric_value / 100
                        
                        # Apply background color based on original value
                        if numeric_value >= 0:
                            cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")  # Light green
                        else:
                            cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
                    except (ValueError, TypeError):
                        cell.value = value
            else:
                # Debug: Print when percentage formatting is NOT applied
                if '%' in clean_metric_name and column_title not in ['Metric_Type', 'Metric', 'Type']:
                    print(f"DEBUG EXCEL: No formatting found for {clean_metric_name} in column {column_title}")
                    print(f"DEBUG EXCEL: Available mappings: {list(config.OPS_METRIC_FORMAT_MAPPING.keys())}")
                    print(f"DEBUG EXCEL: Clean metric name: '{clean_metric_name}'")
                
                # Rest of your formatting logic remains the same
                if column_title in ['Metric_Type', 'Metric', 'Type']:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

                    # # Special handling for Birthday discount cost % of sales row
                    # if is_birthday_cost_current and column_title not in ['Metric', 'Type']:
                    #     cell.number_format = '0.0%'
                    #     continue

                # Apply per-row formatting if provided
                if per_row_formatting:
                    fmt_applied = False
                    for key_col, format_dict in per_row_formatting.items():
                        key_value = data_row.get(key_col)
                        if key_value in format_dict:
                            fmt = format_dict[key_value]
                            if column_title not in merge_columns:
                                if 'number_format' in fmt:
                                    if exception_metric_types and \
                                       any(data_row.get('Metric') == exc_metric and data_row.get('Type') == exc_type for exc_metric, exc_type in exception_metric_types):
                                        current_metric = data_row.get('Metric')
                                        if current_metric == 'Frequency Spend (Rolling 12 months)':
                                            cell.number_format = '#,##0.00'
                                        elif current_metric in ['Loyalty Cost % of Sales', 'Birthday discount cost % of sales']:
                                            cell.number_format = '0.0%'
                                        else:
                                            cell.number_format = fmt['number_format']
                                    else:
                                        cell.number_format = fmt['number_format']
                                if 'font_bold' in fmt and fmt['font_bold']:
                                    cell.font = Font(bold=True)
                            fmt_applied = True
                            break
                    if not fmt_applied:
                        apply_default_formatting(cell, column_title, value, percentage_columns, two_decimal_columns)
                else:
                    apply_default_formatting(cell, column_title, value, percentage_columns, two_decimal_columns)
                    if (is_percent_row == 1 or "%" in str(column_title).strip()) and (col_num>1):
                        try:
                            # Clean the value and convert to float for comparison
                            clean_value = str(cell.value).replace('%', '').replace(',', '')
                            if clean_value and clean_value != 'nan' and clean_value != '':
                                numeric_value = float(clean_value)
                                if numeric_value < 0:
                                    cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                                elif numeric_value > 0:
                                    cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")
                                else:
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        except (ValueError, TypeError):
                            # If conversion fails, use default white background
                            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                    if (str(column_title).strip() == '% RAND GROWTH' or str(column_title).strip() == '% UNIT GROWTH'):
                        try:
                            clean_value = str(cell.value).replace('%', '').replace(',', '')
                            if clean_value and clean_value != 'nan' and clean_value != '':
                                numeric_value = float(clean_value)
                                if numeric_value >= 0:
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        except (ValueError, TypeError):
                            pass

                    if ("Total" in str(data_row[0]).strip()) and (str(data_row[0]).strip() != 'Total Service Items' and str(data_row[0]).strip() != 'Total Retail Items'):
                        cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
                        cell.font = Font(bold=True)
                        
                    if (str(data_row[0]).strip() == 'Total Service Items' or str(data_row[0]).strip() == 'Total Retail Items'):
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    # Apply specific formatting based on column and metric name
                    # check if title in percentage_columns then leave
                    if (col_num > 1 and column_title not in percentage_columns): 
                        if '%' in str(value):
                            if is_number(value):
                                cell.number_format = '0.0%'
                                cell.value = float(str(value).replace('%',''))
                            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                        else:
                            # Apply default number format if it's a number
                            if is_number(value):
                                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                                
                                # Check if this is a frequency metric (should show decimal places)
                                if ('frequency' in str(data_row[0]).lower() and 'rolling' in str(data_row[0]).lower()):
                                    cell.number_format = '#,###,##0.00'
                                else:
                                    cell.number_format = '#,###,##0'
                                
                                try:
                                    cell.value = float(str(value))
                                except (ValueError, TypeError):
                                    cell.value = str(value)
                            else:
                                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Handle merging cells
        if merge_columns:
            for col in merge_columns:
                col_index = df.columns.get_loc(col) + 1
                cell_value = data_row[col]
                merge_info = merge_ranges[col]
                
                if merge_info['value'] == cell_value:
                    # Value is same as previous, extend the merge range
                    merge_info['end_row'] = row
                else:
                    # New value, finalize previous merge if needed
                    if merge_info['value'] is not None and merge_info['start_row'] != row - 1:
                        ws.merge_cells(start_row=merge_info['start_row'], start_column=col_index,
                                     end_row=row - 1, end_column=col_index)
                        merged_cell = ws.cell(row=merge_info['start_row'], column=col_index)
                        merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        merged_cell.font = Font(bold=True)
                        merged_cell.border = create_black_border()
                    # Start new merge range
                    merge_info['value'] = cell_value
                    merge_info['start_row'] = row
                    
        row += 1

    # Finalize any pending merges
    if merge_columns:
        for col in merge_columns:
            merge_info = merge_ranges[col]
            col_index = df.columns.get_loc(col) + 1
            if merge_info['value'] is not None and merge_info['start_row'] != row - 1:
                ws.merge_cells(start_row=merge_info['start_row'], start_column=col_index,
                             end_row=row - 1, end_column=col_index)
                merged_cell = ws.cell(row=merge_info['start_row'], column=col_index)
                merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                merged_cell.font = Font(bold=True)
                merged_cell.border = create_black_border()

    return row

def write_ops_kpi_data(ws, row: int, df: pd.DataFrame, merge_columns: List[str] = None) -> int:
    merge_ranges = {0: {'start_row': row, 'value': None}} if merge_columns else {}

    for idx, data_row in df.iterrows():
        metric_name = str(data_row.iloc[0]).strip()
        clean_metric_name = ' '.join(metric_name.split())

        for col_num, (column_title, value) in enumerate(data_row.items(), 1):
            cell = ws.cell(row=row, column=col_num)
            
            # Identify separator columns and initially set to white fill
            if isinstance(column_title, str) and column_title.startswith('separator_'):
                cell.value = ""
                cell.border = None
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                continue
            
            cell.value = value
            cell.border = create_black_border()


           

            # Handle metric column alignment
            if col_num == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                continue

            # Convert and format data values if applicable
            if isinstance(value, str) and col_num != 1:
                try:
                    is_percentage = clean_metric_name in config.OPS_METRIC_FORMAT_MAPPING and \
                                  config.OPS_METRIC_FORMAT_MAPPING[clean_metric_name].endswith('%')
                    
                    clean_value = value.replace(',', '').replace('%', '')
                    clean_value = float(clean_value)
                    
                    if is_percentage:
                        clean_value = clean_value / 100
                        
                    cell.value = clean_value
                except (ValueError, AttributeError):
                    pass

            if clean_metric_name in config.OPS_METRIC_FORMAT_MAPPING and col_num != 1:
                cell.number_format = config.OPS_METRIC_FORMAT_MAPPING[clean_metric_name]
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

                
            if clean_metric_name in ["Average Basket Size", "Retail Basket Size", "Service Basket Size"] and col_num != 1:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"  # No decimals
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


        # Handle merging cells
        if merge_ranges:
            cell_value = data_row.iloc[0]
            merge_info = merge_ranges[0]
            
            if merge_info['value'] == cell_value:
                merge_info['end_row'] = row
            else:
                if merge_info['value'] is not None and merge_info['start_row'] != row - 1:
                    ws.merge_cells(start_row=merge_info['start_row'], start_column=2,
                                 end_row=row - 1, end_column=2)  # Merge column B
                    merged_cell = ws.cell(row=merge_info['start_row'], column=2)
                    merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    merged_cell.font = Font(bold=True)
                    merged_cell.border = create_black_border()
                merge_info['value'] = cell_value
                merge_info['start_row'] = row
        
        row += 1

    # Finalize merges
    if merge_ranges:
        merge_info = merge_ranges[0]
        if merge_info['value'] is not None and merge_info['start_row'] != row - 1:
            ws.merge_cells(start_row=merge_info['start_row'], start_column=1,
                         end_row=row - 1, end_column=1)
            merged_cell = ws.cell(row=merge_info['start_row'], column=1)
            merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            merged_cell.font = Font(bold=True)
            merged_cell.border = create_black_border()

    # Apply a final white fill to separator columns after all other formatting
    for row_idx in range(row - len(df), row):
        for col_num, column_title in enumerate(df.columns, 1):
            if isinstance(column_title, str) and column_title.startswith('separator_'):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


    return row

def write_clicks_loyalty_data(ws, row: int, df: pd.DataFrame, merge_columns: List[str] = None) -> int:
    merge_ranges = {0: {'start_row': row, 'value': None}} if merge_columns else {}

    # Define the exact section headers that should be grey
    section_headers = {
        'Sales',             # Only the main section header "Sales"
        'Transactions',      # Only the main section header "Transactions"
        'Guests',           # Only the main section header "Guest Count"
        'Basket size'        # Only the main section header "Basket Size"
    }

    for idx, data_row in df.iterrows():
        # Get the metric name from the first column
        metric_name = str(data_row.iloc[0]).strip()
        # Remove any leading/trailing whitespace and special characters
        clean_metric_name = ' '.join(metric_name.split())

        # Check if this row should be grey (exact match for section headers only)
        is_section_header = clean_metric_name in section_headers

        for col_num, (column_title, value) in enumerate(data_row.items(), 2):  # Start from column B (2)
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.border = create_black_border()

            # Apply grey background for section headers
            if is_section_header:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.font = Font(bold=True)

            # Handle the metric column (first column - now column B)
            if col_num == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                continue

            # Check if this is the Growth column (case-insensitive)
            is_growth_column = str(column_title).lower() == 'growth'
            
            # Define percentage metrics (including those not in config mapping)
            percentage_metrics = {
                "Clicks loyalty sales over total",
                "Clicks loyalty transactions over total", 
                "Clicks loyalty basket size over total",
                "New clicks loyalty guest over new guests",
                "Clicks loyalty new guests over total unique guests"
            }

            # Convert string values to numbers if possible
            if isinstance(value, str) and col_num != 2:
                try:
                    # Check if the metric requires percentage formatting
                    is_percentage = is_growth_column or (
                        clean_metric_name in config.CLICKS_LOYALTY_FORMAT_MAPPING and 
                        config.CLICKS_LOYALTY_FORMAT_MAPPING[clean_metric_name].endswith('%')
                    ) or clean_metric_name in percentage_metrics
                    
                    # Remove commas and '%' signs, then convert to float
                    clean_value = value.replace(',', '').replace('%', '')
                    clean_value = float(clean_value)
                    
                    # If it's a percentage, divide by 100
                    if is_percentage:
                        clean_value = clean_value / 100
                        
                    cell.value = clean_value
                except (ValueError, AttributeError):
                    pass
            elif isinstance(value, (int, float)) and col_num != 2:
                # Handle numeric values directly
                cell.value = value

            # Apply specific formatting based on column and metric name
            if is_growth_column:
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                
                # Apply background color based on positive/negative value
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 0:
                        cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")  # Light green
                    else:
                        cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
                        
            elif clean_metric_name in config.CLICKS_LOYALTY_FORMAT_MAPPING and col_num != 2:
                cell.number_format = config.CLICKS_LOYALTY_FORMAT_MAPPING[clean_metric_name]
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                
                # Apply background color for percentage columns
                if config.CLICKS_LOYALTY_FORMAT_MAPPING[clean_metric_name].endswith('%') and isinstance(cell.value, (int, float)):
                    if cell.value >= 0:
                        cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")  # Light green
                    else:
                        cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
            elif clean_metric_name in percentage_metrics and col_num != 2:
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                
                # Apply background color for percentage columns
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 0:
                        cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")  # Light green
                    else:
                        cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                # Apply default number format if it's a number
                if isinstance(cell.value, (int, float)):
                    # Check if this is a frequency metric (should show decimal places)
                    if ('frequency' in clean_metric_name.lower() and 'rolling' in clean_metric_name.lower()):
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0'

        # Handle merging cells
        if merge_ranges:
            cell_value = data_row.iloc[0]
            merge_info = merge_ranges[0]
            
            if merge_info['value'] == cell_value:
                merge_info['end_row'] = row
            else:
                if merge_info['value'] is not None and merge_info['start_row'] != row - 1:
                    ws.merge_cells(start_row=merge_info['start_row'], start_column=2,
                                 end_row=row - 1, end_column=2)  # Merge column B
                    merged_cell = ws.cell(row=merge_info['start_row'], column=2)
                    merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    merged_cell.font = Font(bold=True)
                    merged_cell.border = create_black_border()
                merge_info['value'] = cell_value
                merge_info['start_row'] = row
        
        row += 1

    # Finalize any pending merges
    if merge_ranges:
        merge_info = merge_ranges[0]
        if merge_info['value'] is not None and merge_info['start_row'] != row - 1:
            ws.merge_cells(start_row=merge_info['start_row'], start_column=1,
                         end_row=row - 1, end_column=1)
            merged_cell = ws.cell(row=merge_info['start_row'], column=1)
            merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            merged_cell.font = Font(bold=True)
            merged_cell.border = create_black_border()

    # Column widths are set globally in the write_data function

    return row
def apply_default_formatting(cell, column_title, value, percentage_columns, two_decimal_columns):
    # Apply default formatting
    if column_title in percentage_columns:
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell.number_format = "0.0%"  # Apply percentage format with 1 decimal place
    elif isinstance(value, (int, float)):
        if column_title in two_decimal_columns:
            cell.number_format = '#,##0.00'
        else:
            cell.number_format = '#,##0'  # Zero decimal places
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def bold_total_rows(ws, df: pd.DataFrame, start_row: int,
                    percentage_columns: List[str], two_decimal_columns: List[str],
                    first_column: int = 1, current_subheader: str = None):
    # Apply special category highlighting for all Ops KPI sections
    if current_subheader and current_subheader.startswith("Ops KPI"):
        highlight_categories = {'Sales', 'Transactions and Basket Size', 'Other KPIs'}
        
        # Iterate over rows using DataFrame index to avoid attribute access errors
        for i, row_data in enumerate(df.itertuples(index=False), start=start_row):
            metric_type = str(row_data[first_column - 1]).strip()

            if metric_type in highlight_categories:
                for col_num in range(1, len(df.columns) + 1):
                    column_title = df.columns[col_num - 1]
                    cell = ws.cell(row=i, column=col_num)
                    
                    # Check if this is a separator column
                    if isinstance(column_title, str) and column_title.startswith('separator_'):
                        # Ensure separator columns remain white
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    else:
                        # Apply gray fill
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    
                    if column_title in percentage_columns:
                        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                        cell.number_format = "0.0%"
                    elif isinstance(cell.value, (int, float)):
                        if column_title in two_decimal_columns:
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    else:
        # Original total/group logic for other sections
        total_keywords = {'total', 'group'}
        
        for i, row_data in enumerate(df.itertuples(index=False), start=start_row):
            cell_value = str(row_data[first_column - 1]).strip().lower()

            if cell_value in total_keywords:
                for col_num in range(1, len(df.columns) + 1):
                    column_title = df.columns[col_num - 1]
                    cell = ws.cell(row=i, column=col_num)
                    
                    # Check if this is a separator column
                    if isinstance(column_title, str) and column_title.startswith('separator_'):
                        # Ensure separator columns remain white
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    else:
                        # Apply gray fill
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    
                    if column_title in percentage_columns:
                        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                        cell.number_format = "0.0%"
                    elif isinstance(cell.value, (int, float)):
                        if column_title in two_decimal_columns:
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def export_combined_excel_from_template(
    all_dataframes: List[pd.DataFrame],
    subheaders: List[str],
    start_date: str,
    end_date: str,
    template_path: str = "clicks_scorecard_template.xlsx",
    columns_with_two_decimals: Dict[str, List[str]] = None,
    fiscal_year_start: str = None,
    execution_datetime: str = None
) -> bytes:
    # Pre-process Ops KPI dataframes
    # all_dataframes, subheaders = combine_ops_kpi_dataframes(all_dataframes, subheaders)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_full_path = os.path.join(script_dir, template_path)

    wb, ws = load_template(template_full_path)

    # Add comprehensive report information on row 2, merged across 8 columns
    current_datetime = datetime.now()
    
    # Use the reporting week dates passed as parameters (not current week)
    week_start = datetime.strptime(start_date, '%Y-%m-%d')
    week_end = datetime.strptime(end_date, '%Y-%m-%d')
    
    # Calculate month start and financial year start (September 1st)
    month_start = current_datetime.replace(day=1)
    
    # Financial year start follows the same logic as YTD calculations
    if current_datetime.month > 9 or (current_datetime.month == 9 and current_datetime.day > 1):
        financial_year_start = current_datetime.replace(month=9, day=1)
    else:
        financial_year_start = current_datetime.replace(year=current_datetime.year - 1, month=9, day=1)
    
    # Create comprehensive report text with fiscal year and execution time
    if fiscal_year_start and execution_datetime:
        report_text = f"Weekly Performance ({week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')}), Fiscal Year Start: {fiscal_year_start}, Execution Date & Time: {execution_datetime}"
    else:
        # Fallback to calculating fiscal year and execution time if not provided
        fiscal_year_start_str = fiscal_year_start.strftime('%Y-%m-%d') if fiscal_year_start else financial_year_start.strftime('%Y-%m-%d')
        execution_datetime = execution_datetime if execution_datetime else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        report_text = f"Weekly Performance ({week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')}), Fiscal Year Start: {fiscal_year_start_str}, Execution Date & Time: {execution_datetime}"
    
    # Write report info to cell B2
    ws.cell(row=2, column=2, value=report_text)
    ws.cell(row=2, column=2).font = Font(bold=True, size=11)
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=2).fill = PatternFill(start_color="F0F2F6", end_color="F0F2F6", fill_type="solid")
    
    # Merge cells B2:J2 (9 columns)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=10)
    
    # Leave one blank row (row 3), then start tables from row 4
    exception_subheaders = ["Total Lines Service", "Total Lines Retail"]

    current_row = 4

    for df, subheader in zip(all_dataframes, subheaders):
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue

        # Identify percentage columns
        percentage_columns = [col for col in df.columns if '%' in col]

        # Process percentage columns to ensure they are numeric and values are between -1 and 1
        for col in percentage_columns:
            # Remove any '%' signs from the data
            df[col] = df[col].replace('%', '', regex=True)
            # Convert to float
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            # If absolute values are >=1, divide by 100 to convert to fractions
            df[col] = df[col] / 100

        # Get list of columns that need two decimal places for this subheader
        if columns_with_two_decimals and subheader in columns_with_two_decimals:
            two_decimal_columns = columns_with_two_decimals[subheader]
        else:
            two_decimal_columns = []

        # Write subheader as a bold title
        # Add merging here -->
        if subheader=='Total Service Items' or subheader=='Total Retail Items' :
            ws.cell(row=current_row, column=2, value=subheader)
        elif subheader=='Retail Top 20 Items' or subheader=='Service Top 20 Items':
            weekly_header='Weekly Performance (' + start_date + ' to ' + end_date + ')'
            ws.cell(row=current_row, column=2, value=weekly_header)
            ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='left')
        else:
            ws.cell(row=current_row, column=2, value=subheader)
            ws.cell(row=current_row, column=2).font = Font(bold=True, size=12, color='FFFFFF')
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            
            # For regional tables, merge across all columns (8 columns total - one less)
            if subheader in ["Last Week Sales by Region", "Month to Date by Region", "Year to Date by Region", 
                           "Transaction Count: Last Week by Region", "Basket size: Last Week by Region", "Week to Date by Region", "Basket Size by Region", "Transaction Count by Region"]:
                ws.merge_cells(start_row=current_row, start_column=2,
                                end_row=current_row, end_column=9)  # Merge across 8 columns (one less)
            # For Clicks Loyalty KPI, merge across all columns (typically 4 columns: Metric Type, Current, Previous, Growth)
            elif subheader == "Clicks Loyalty KPI":
                ws.merge_cells(start_row=current_row, start_column=2,
                                end_row=current_row, end_column=5)  # Merge across all 4 columns
            # For Daily Sales Last Week, merge across all columns (typically 7 columns)
            elif subheader.startswith("Daily Sales Last Week"):
                ws.merge_cells(start_row=current_row, start_column=2,
                                end_row=current_row, end_column=8)  # Merge across all 7 columns
            # For Loyalty KPI, merge across 9 columns (Metric, Type, Green, Blue, Silver, Gold, Total Loyalty, Non-loyalty, Total)
            elif subheader == "Loyalty KPI":
                ws.merge_cells(start_row=current_row, start_column=2,
                                end_row=current_row, end_column=10)  # Merge across 9 columns
            # For Basket Size Total, merge across 4 columns (ACT_BUD, WTD, MTD, YTD)
            elif subheader == "Basket Size Total":
                ws.merge_cells(start_row=current_row, start_column=2,
                                end_row=current_row, end_column=5)  # Merge across 4 columns
            # For Basket Size by Region, merge across 19 columns
            elif subheader == "Basket Size by Region":
                ws.merge_cells(start_row=current_row, start_column=2,
                                end_row=current_row, end_column=20)  # Merge across 19 columns
            # For Top 10 and ICU Stores, merge across 10 columns
            elif subheader in ["Top 10 Stores - Current Sales", "ICU Stores - Current Sales"]:
                ws.merge_cells(start_row=current_row, start_column=2,
                                end_row=current_row, end_column=11)  # Merge across 10 columns
            else:
                ws.merge_cells(start_row=current_row, start_column=2,
                                end_row=current_row, end_column=5)  # Default merge for other tables

        current_row += 1

        # Filter out RAW columns for Loyalty KPI to avoid showing them in headers
        if subheader == "Loyalty KPI":
            display_columns = [col for col in df.columns if not col.endswith('_RAW')]
            write_headers(ws, current_row, display_columns,
                          exception_subheaders, subheader)
        else:
            write_headers(ws, current_row, df.columns.tolist(),
                          exception_subheaders, subheader)
        current_row += 1

        data_start_row = current_row

        # Choose appropriate write function based on section
        if subheader.startswith("Ops KPI"):
            current_row = write_ops_kpi_data(ws, current_row, df, merge_columns=[0])
        elif subheader.startswith("Daily Sales Last Week"):
            current_row = write_daily_sales_data(ws, current_row, df)
        elif subheader == "Clicks Loyalty KPI":
            # Add period information row for Clicks Loyalty KPI
            period_info_row = current_row
            current_row += 1
            
            # Calculate previous period dates
            # Parse the start and end dates
            start_dt = pd.to_datetime(start_date)
            end_dt = pd.to_datetime(end_date)
            
            # Calculate previous period (same week last year)
            prev_start_dt = start_dt - pd.DateOffset(years=1)
            prev_end_dt = end_dt - pd.DateOffset(years=1)
            
            # Format dates
            current_start_str = start_dt.strftime('%Y-%m-%d')
            current_end_str = end_dt.strftime('%Y-%m-%d')
            prev_start_str = prev_start_dt.strftime('%Y-%m-%d')
            prev_end_str = prev_end_dt.strftime('%Y-%m-%d')
            
            # Create period information text
            period_text = f"Current Week: {current_start_str} to {current_end_str} | Previous Week (Same Week Last Year): {prev_start_str} to {prev_end_str}"
            
            # Write period information to Excel
            period_cell = ws.cell(row=period_info_row, column=2, value=period_text)
            period_cell.font = Font(bold=True, size=10)
            period_cell.alignment = Alignment(horizontal="left", vertical="center")
            period_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue background
            
            # Merge across all columns for the period information
            ws.merge_cells(start_row=period_info_row, start_column=2, end_row=period_info_row, end_column=5)
            
            current_row = write_clicks_loyalty_data(ws, current_row, df, merge_columns=[0])
        elif subheader == "Loyalty KPI":
            current_row = write_loyalty_kpi_data(ws, current_row, df)
        elif subheader in ["Basket Size Total", "Basket Size by Region"]:
            # Use dedicated function for Basket Size tables with proper percentage formatting
            current_row = write_basket_size_data(ws, current_row, df)
        elif subheader in ["Top 10 Stores - Current Sales", "ICU Stores - Current Sales"]:
            # Use dedicated function for Top 10 and ICU stores with proper total row calculations
            current_row = write_stores_data(ws, current_row, df)
        else:
            # Original handling for other sections
            current_row = write_data(ws, current_row, df, percentage_columns, two_decimal_columns, subheader=subheader)

        data_end_row = current_row - 1

        # Adjust column widths
        for col_num, column_title in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in ws[col_letter]
                if cell.row >= data_start_row - 2
            )
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Pass percentage_columns and two_decimal_columns to bold_total_rows
        bold_total_rows(ws, df, data_start_row,
                       percentage_columns, two_decimal_columns,
                       first_column=1, current_subheader=subheader)

        current_row += 1

    # Set the first column width to 25
    ws.column_dimensions['A'].width = 4.71
    ws.column_dimensions['B'].width = 45
    fixed_width = 15.00
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter != 'A' and col_letter != 'B':
            ws.column_dimensions[col_letter].width = fixed_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()


def write_daily_sales_data(ws, start_row: int, df: pd.DataFrame) -> int:
    """
    Write daily sales data to Excel worksheet with proper formatting.
    """
    current_row = start_row
    
    for idx, row in df.iterrows():
        # Write each column
        for col_idx, (col_name, value) in enumerate(row.items()):
            cell = ws.cell(row=current_row, column=col_idx + 2, value=value)
            
            # Apply black borders to all cells
            cell.border = create_black_border()
            
            # Center align dates in DATES_CY column
            if col_name == 'DATES_CY':
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Format based on row type
            if row['SALES'] in ['Total', 'Weekend trading (Fri/Sat/Sun)', 'AVG Sales']:
                # Summary rows - bold text
                cell.font = Font(bold=True)
                
                # Color coding for summary rows
                if row['SALES'] == 'Total':
                    # Total row - light blue background
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                elif row['SALES'] == 'Weekend trading (Fri/Sat/Sun)':
                    # Weekend row - light green background
                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                elif row['SALES'] == 'AVG Sales':
                    # Average row - light yellow background
                    cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                
                # Apply number formatting to summary rows - CRITICAL FIX!
                if col_name in ['SALES_TY', 'SALES_LY', 'SALES_BUDGET']:
                    # Numbers with thousand separators and no decimals (same as Total row)
                    cell.number_format = '#,##0'
                elif col_name in ['GROWTH_PCT', 'BUDGET_PCT']:
                    # Apply percentage formatting to summary rows - special handling
                    # Always use the raw decimal value and let Excel format it
                    if isinstance(value, (int, float)):
                        print(f"DEBUG Excel - Using numeric value: {value} for {col_name}")
                        cell.value = value  # Use raw decimal value
                        cell.number_format = '0.0%'
                    elif isinstance(value, str) and '%' in str(value):
                        # Convert formatted string back to decimal
                        numeric_value = float(str(value).replace('%', '')) / 100
                        print(f"DEBUG Excel - Converting string '{value}' to {numeric_value} for {col_name}")
                        cell.value = numeric_value
                        cell.number_format = '0.0%'
                    else:
                        print(f"DEBUG Excel - Using default formatting for {col_name}")
                        cell.number_format = '0.0%'
                
                # Color code only GROWTH_PCT column (remove colors from BUDGET_PCT)
                if col_name in ['GROWTH_PCT']:
                    try:
                        # Use RAW value for color coding if available
                        raw_col_name = f'{col_name}_RAW'
                        if raw_col_name in row.index and pd.notna(row[raw_col_name]):
                            pct_value = float(row[raw_col_name]) * 100  # Convert decimal to percentage
                        elif isinstance(value, (int, float)):
                            pct_value = float(value) * 100  # Convert decimal to percentage
                        elif isinstance(value, str) and '%' in str(value):
                            pct_value = float(str(value).replace('%', ''))
                        else:
                            pct_value = 0
                        
                        if pct_value >= 0:
                            cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")
                        else:
                            # Softer red background
                            cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                    except (ValueError, TypeError) as e:
                        print(f"DEBUG Excel - Error parsing GROWTH_PCT '{value}': {e}")
                        pass
                elif col_name in ['BUDGET_PCT']:
                    # Make BUDGET_PCT grey like the rest of the Total row
                    if row['SALES'] == 'Total':
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            else:
                # Regular daily rows - apply number formatting
                if col_name in ['SALES_TY', 'SALES_LY', 'SALES_BUDGET']:
                    # Numbers with thousand separators and no decimals
                    cell.number_format = '#,##0'
                elif col_name in ['GROWTH_PCT', 'BUDGET_PCT']:
                    # For daily rows, handle both raw decimals and formatted strings
                    if isinstance(value, (int, float)):
                        cell.value = value  # Use raw decimal value
                        cell.number_format = '0.0%'
                    elif isinstance(value, str) and '%' in str(value):
                        # Convert formatted string back to decimal
                        numeric_value = float(str(value).replace('%', '')) / 100
                        cell.value = numeric_value
                        cell.number_format = '0.0%'
                    else:
                        cell.number_format = '0.0%'
                    
                    # Color code only GROWTH_PCT (remove colors from BUDGET_PCT)
                    if col_name == 'GROWTH_PCT':
                        try:
                            # Handle both raw decimals and formatted strings for color coding
                            if isinstance(value, (int, float)):
                                pct_value = float(value) * 100  # Convert decimal to percentage
                            elif isinstance(value, str) and '%' in str(value):
                                pct_value = float(str(value).replace('%', ''))
                            else:
                                pct_value = 0
                            
                            if pct_value >= 0:
                                cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")
                            else:
                                # Softer red background
                                cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                        except (ValueError, TypeError):
                            pass
        
        current_row += 1
    
    return current_row

def is_number(value):
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False



def write_loyalty_kpi_data(ws, start_row: int, df: pd.DataFrame) -> int:
    """
    Writes the Loyalty KPI data to Excel with proper formatting:
    - Formats growth percentages with background colors (green/red)
    - Formats numeric values with thousand separators
    - Uses raw decimal values for proper Excel formatting
    - Excludes helper columns (_RAW columns) from display
    """
    current_row = start_row
    
    # Define the columns that contain growth percentages (exclude _RAW helper columns)
    growth_columns = ['Green', 'Blue', 'Silver', 'Gold', 'Total Loyalty', 'Non-loyalty', 'Total']
    
    # Filter out helper columns (_RAW columns) for display
    display_columns = [col for col in df.columns if not col.endswith('_RAW')]
    
    for row_idx, row in df.iterrows():
        for col_idx, col_name in enumerate(display_columns, 2):  # Start from column 2 (B) instead of 1 (A)
            cell = ws.cell(row=current_row, column=col_idx)
            value = row[col_name]
            
            # Apply black borders to all cells
            cell.border = create_black_border()
            
            if col_name == 'Metric':
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(bold=True)
            elif col_name == 'Type':
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(size=8)
            elif col_name in growth_columns:
                # Check if this is a "Growth period on period" row
                if row['Type'] == 'Growth period on period':
                    # Use RAW column if available for proper Excel formatting
                    raw_col_name = f'{col_name}_RAW'
                    if raw_col_name in row.index and pd.notna(row[raw_col_name]):
                        cell.value = row[raw_col_name]  # Use raw decimal value
                        cell.number_format = '0.0%'
                        
                        # Apply background color based on value
                        pct_value = float(row[raw_col_name]) * 100  # Convert decimal to percentage
                        if pct_value >= 0:
                            cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")  # Light green
                        else:
                            cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
                    else:
                        # Fallback to formatted string
                        cell.value = value
                        cell.number_format = '0.0%'
                elif row['Type'] == 'Current':
                    # Check if this is a percentage metric (contains '%' in the Metric name)
                    metric_name = str(row['Metric']).lower()
                    is_percentage_metric = '%' in metric_name
                    is_frequency_metric = 'freq' in metric_name or 'frequency' in metric_name
                    
                    if is_percentage_metric:
                        # Use RAW column if available for proper Excel formatting (same as Growth period on period)
                        raw_col_name = f'{col_name}_RAW'
                        if raw_col_name in row.index and pd.notna(row[raw_col_name]):
                            cell.value = row[raw_col_name]  # Use raw decimal value
                            cell.number_format = '0.0%'
                        else:
                            # Fallback to formatted string
                            cell.value = value
                            cell.number_format = '0.0%'
                    elif pd.notna(value) and value != 0:
                        # Convert to numeric properly to avoid Excel text warnings
                        try:
                            numeric_value = float(str(value).replace(',', '').replace(' ', ''))
                            cell.value = numeric_value
                        except (ValueError, TypeError):
                            cell.value = 0
                        
                        if is_frequency_metric:
                            # Format frequency values with 1 decimal place
                            cell.number_format = '#,##0.0'
                        else:
                            # Format other numeric values with thousand separators and no decimals
                            cell.number_format = '#,##0'
                    else:
                        cell.value = 0
                        if is_frequency_metric:
                            cell.number_format = '#,##0.0'
                        else:
                            cell.number_format = '#,##0'
                elif row['Type'] == 'Contribution to total':
                    # Format contribution percentages
                    raw_col_name = f'{col_name}_RAW'
                    if raw_col_name in row.index and pd.notna(row[raw_col_name]):
                        cell.value = row[raw_col_name]  # Use raw decimal value
                        cell.number_format = '0.0%'
                    else:
                        cell.value = value
                        cell.number_format = '0.0%'
                
                cell.alignment = Alignment(horizontal="right", vertical="center")
        
        current_row += 1
    
    return current_row

def write_basket_size_data(ws, row: int, df: pd.DataFrame) -> int:
    """Dedicated function for Basket Size tables with proper percentage formatting"""
    
    for idx, data_row in df.iterrows():
        # Get the metric name from the first column (act_bud)
        metric_name = str(data_row.iloc[0]).strip()
        
        for col_num, (column_title, value) in enumerate(data_row.items(), 1):  # Start from column 1 (headers already offset)
            cell = ws.cell(row=row, column=col_num+1)  # Add +1 to match header offset
            cell.value = value
            cell.border = create_black_border()
            
            # Check if this is a percentage row
            if '%' in metric_name and col_num > 1:  # Skip the first column (act_bud)
                if pd.notna(value) and value != '':
                    try:
                        numeric_value = float(value)
                        # Convert from percentage (1.27) to decimal (0.0127) for Excel's 0.0% format
                        cell.value = numeric_value / 100
                        cell.number_format = '0.0%'
                        
                        # Apply background color based on original value
                        if numeric_value >= 0:
                            cell.fill = PatternFill(start_color="C6E2B7", end_color="C6E2B7", fill_type="solid")  # Light green
                        else:
                            cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
                    except (ValueError, TypeError):
                        cell.value = value
                        cell.number_format = '0.0%'
                else:
                    cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                # Regular formatting for non-percentage rows
                if col_num == 1:  # First column (act_bud)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
        
        row += 1
    
    return row

def write_stores_data(ws, row: int, df: pd.DataFrame) -> int:
    """Dedicated function for Top 10 and ICU stores tables with proper total row calculations"""
    
    # Calculate proper totals from individual rows (excluding Total row)
    non_total_rows = df[df.iloc[:, 0] != "Total"]
    
    for idx, data_row in df.iterrows():
        is_total_row = str(data_row.iloc[0]).strip() == "Total"
        
        for col_num, (column_title, value) in enumerate(data_row.items(), 1):  # Start from column 1 (headers already offset)
            cell = ws.cell(row=row, column=col_num+1)  # Add +1 to match header offset
            cell.border = create_black_border()
            
            # Apply grey background to Total row starting from column B (col_num > 1)
            if is_total_row and col_num > 1:
                cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
                cell.font = Font(bold=True)
            
            if is_total_row:
                # Special handling for Total row
                if col_num == 1:  # First column (CENTER)
                    cell.value = "Total"
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif column_title in []:  # No percentage columns left after removal
                    # For percentage columns in total row, show 100.0%
                    cell.value = 1.0  # 1.0 for Excel's 0.0% format = 100.0%
                    cell.number_format = '0.0%'
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                elif column_title == "TOTAL REVENUE CURRENT":
                    # Special handling for TOTAL REVENUE CURRENT - sum SERVICE + RETAIL
                    try:
                        service_values = pd.to_numeric(non_total_rows["SERVICE REVENUE CURRENT"], errors='coerce')
                        retail_values = pd.to_numeric(non_total_rows["RETAIL REVENUE CURRENT"], errors='coerce')
                        total_revenue = service_values.sum() + retail_values.sum()
                        cell.value = total_revenue
                        cell.number_format = '#,##0'
                    except:
                        cell.value = value
                        cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                elif column_title == "TOTAL REVENUE PREVIOUS":
                    # Special handling for TOTAL REVENUE PREVIOUS - sum SERVICE + RETAIL
                    try:
                        service_values = pd.to_numeric(non_total_rows["SERVICE REVENUE PREVIOUS"], errors='coerce')
                        retail_values = pd.to_numeric(non_total_rows["RETAIL REVENUE PREVIOUS"], errors='coerce')
                        total_revenue = service_values.sum() + retail_values.sum()
                        cell.value = total_revenue
                        cell.number_format = '#,##0'
                    except:
                        cell.value = value
                        cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                elif column_title == "TOTAL BUDGET":
                    # Special handling for TOTAL BUDGET - sum SERVICE + RETAIL
                    try:
                        service_values = pd.to_numeric(non_total_rows["SERVICE BUDGET"], errors='coerce')
                        retail_values = pd.to_numeric(non_total_rows["RETAIL BUDGET"], errors='coerce')
                        total_budget = service_values.sum() + retail_values.sum()
                        cell.value = total_budget
                        cell.number_format = '#,##0'
                    except:
                        cell.value = value
                        cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                else:
                    # For other revenue/budget columns, calculate proper totals
                    if column_title in non_total_rows.columns:
                        try:
                            # Sum only numeric values from non-total rows
                            numeric_values = pd.to_numeric(non_total_rows[column_title], errors='coerce')
                            total_value = numeric_values.sum()
                            cell.value = total_value
                            cell.number_format = '#,##0'
                        except:
                            cell.value = value
                            cell.number_format = '#,##0'
                    else:
                        cell.value = value
                        cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                # Regular row formatting
                cell.value = value
                if col_num == 1:  # First column (CENTER)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif column_title in []:  # No percentage columns left after removal
                    # Percentage columns
                    if pd.notna(value) and value != '':
                        try:
                            if isinstance(value, str) and value.endswith('%'):
                                # Already formatted percentage
                                numeric_value = float(value.replace('%', ''))
                                cell.value = numeric_value / 100
                                cell.number_format = '0.0%'
                            else:
                                numeric_value = float(value)
                                cell.value = numeric_value / 100
                                cell.number_format = '0.0%'
                        except (ValueError, TypeError):
                            cell.value = value
                            cell.number_format = '0.0%'
                    else:
                        cell.number_format = '0.0%'
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                else:
                    # Numeric columns
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
        
        row += 1
    
    return row