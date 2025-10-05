'''
This file contains the functions to export the data to an Excel file.
'''
import streamlit as st
from io import BytesIO
import config
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = "group_sales_template.xlsx"

def write_dataframe_to_worksheet(ws, df, start_row=4, start_col=1):
    """
    Writes the DataFrame to the worksheet starting from the specified row and column.
    Returns the total number of rows written.
    """
    data_rows = list(dataframe_to_rows(df, index=False, header=False))
    for idx, row in enumerate(data_rows):
        r_idx = start_row + idx
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)
    return len(data_rows)

def apply_number_formatting(ws, df, start_row=4, start_col=1):
    """
    Applies number formatting to numeric cells in the worksheet.
    """
    numeric_columns = df.select_dtypes(include=['number']).columns
    numeric_col_indices = [df.columns.get_loc(col) + start_col for col in numeric_columns]

    total_rows = ws.max_row
    for row in ws.iter_rows(min_row=start_row, max_row=total_rows, min_col=start_col, max_col=ws.max_column):
        for cell in row:
            if cell.column in numeric_col_indices:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

def apply_conditional_formatting(ws, df, start_row=4, start_col=1):
    """
    Applies uniform blue (#A6C9EC) conditional formatting for total rows.
    Skips empty columns.
    """
    # Define fill and font
    blue_fill = PatternFill(start_color='A6C9EC', end_color='A6C9EC', fill_type='solid')
    bold_font = Font(bold=True)

    # Identify empty columns
    empty_columns = ['Empty1', 'Empty2']
    empty_col_indices = []
    for col_name in empty_columns:
        try:
            empty_col_indices.append(df.columns.get_loc(col_name) + start_col)
        except KeyError:
            continue

    total_rows = ws.max_row
    for row_idx in range(start_row, total_rows + 1):
        brand_cell = ws.cell(row=row_idx, column=1)
        brand_value = brand_cell.value or ''

        # Convert to string and uppercase for case-insensitive comparison
        brand_value_upper = str(brand_value).upper()

        # Check if it's a total row
        if 'TOTAL BY REGION' in brand_value_upper or 'TOTAL' in brand_value_upper:
            # Apply formatting to all columns except empty ones
            for col_idx in range(1, 26):  # A to Z
                if col_idx not in empty_col_indices:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = blue_fill
                    cell.font = bold_font

def apply_performance_formatting(ws, df, target_columns, start_row=4, start_col=1, threshold=100):
    """
    Applies performance-based formatting (green/red) to multiple specified columns.
    Skips total rows as they have their own formatting.
    """
    # Define fills and fonts for performance
    good_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    bad_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    good_font = Font(color='006100')
    bad_font = Font(color='9C0006')

    # Get column indices for all target columns
    column_indices = {}
    for col_name in target_columns:
        try:
            column_indices[col_name] = df.columns.get_loc(col_name) + start_col
        except KeyError:
            st.error(f"Column '{col_name}' not found in DataFrame")

    total_rows = ws.max_row
    for row_idx in range(start_row, total_rows + 1):
        # Check if it's a total row (skip if it is)
        brand_cell = ws.cell(row=row_idx, column=1)
        if 'TOTAL' in str(brand_cell.value or '').upper():
            continue

        # Apply formatting to each target column
        for col_name, col_index in column_indices.items():
            cell = ws.cell(row=row_idx, column=col_index)
            
            try:
                value = float(cell.value if cell.value is not None else 0)
                if value >= threshold:
                    cell.fill = good_fill
                    cell.font = good_font
                else:
                    cell.fill = bad_fill
                    cell.font = bad_font
            except (ValueError, TypeError):
                continue
def merge_brand_cells(ws, start_row=4):
    """
    Merges cells in the BRAND column where the value is the same, excluding rows where the value contains 'TOTAL'.
    Sets the alignment of the merged cells to top-left.
    """
    total_rows = ws.max_row
    merge_start_row = None
    previous_brand = None

    for row_idx in range(start_row, total_rows + 1):
        brand_cell = ws.cell(row=row_idx, column=1)
        brand_value = brand_cell.value or ''
        is_total_row = 'TOTAL' in str(brand_value).upper()

        if not is_total_row:
            if brand_value == previous_brand:
                # Same brand as previous row, continue the merge
                brand_cell.value = None  # Blank out the cell
            else:
                # New brand encountered, merge the previous group if exists
                if merge_start_row and merge_start_row != row_idx - 1:
                    # Merge cells from merge_start_row to previous row
                    ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=row_idx - 1, end_column=1)
                    # Set alignment for the merged cell
                    top_cell = ws.cell(row=merge_start_row, column=1)
                    top_cell.alignment = Alignment(horizontal='left', vertical='top')
                merge_start_row = row_idx
                previous_brand = brand_value
        else:
            # For 'TOTAL' rows, merge the previous group if exists
            if merge_start_row and merge_start_row != row_idx - 1:
                ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=row_idx - 1, end_column=1)
                # Set alignment for the merged cell
                top_cell = ws.cell(row=merge_start_row, column=1)
                top_cell.alignment = Alignment(horizontal='left', vertical='top')
            merge_start_row = None
            previous_brand = None

    # Merge the last group if needed
    if merge_start_row and merge_start_row != total_rows:
        ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=total_rows, end_column=1)
        # Set alignment for the merged cell
        top_cell = ws.cell(row=merge_start_row, column=1)
        top_cell.alignment = Alignment(horizontal='left', vertical='top')

def format_header(ws, sheet_name, start_date, end_date):
    """
    Formats the header row with the worksheet name and date range.
    """
    # Merge cells from A1 to D1 (adjust the range if needed)
    ws.merge_cells('A2:E2')
    formatted_sheet_name = sheet_name.replace('_', ' ')
    # Write the worksheet name and date range into the merged cell
    ws['A2'] = f"{formatted_sheet_name}: {start_date} to {end_date}"

    # Set alignment to left
    ws['A2'].alignment = Alignment(horizontal='left')

    # Set font size to 16 and make it bold
    ws['A2'].font = Font(size=16, bold=True)

def apply_gridlines(ws, df, start_row=4, start_col=1):
    """
    Applies gridlines with bold outlines around specific column groupings.
    """
    # Define border styles
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Get the last row with data
    total_rows = len(df.index)
    end_row = start_row + total_rows - 1
    
    # Updated column groups to include Store Status
    groups = [
        (1, 5),      # First 5 columns (including Store Status)
        (7, 9),      # After Empty1, next 3
        (10, 12),    # Next 3
        (13, 15),    # Next 3
        (17, 19),    # After Empty2, next 3
        (20, 22),    # Next 3
        (23, 25)     # Final 3
    ]


    # Apply thin borders to all cells in the data range
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=25):
        for cell in row:
            cell.border = thin_border

    # Apply thick borders to each group
    for group_start, group_end in groups:
        # Top border for first row of group
        for col in range(group_start, group_end + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=thick,
                bottom=cell.border.bottom
            )

        # Bottom border for last row of group
        for col in range(group_start, group_end + 1):
            cell = ws.cell(row=end_row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=thick
            )

        # Left and right borders for each group
        for row in range(start_row, end_row + 1):
            # Left border
            left_cell = ws.cell(row=row, column=group_start)
            left_cell.border = Border(
                left=thick,
                right=left_cell.border.right,
                top=left_cell.border.top,
                bottom=left_cell.border.bottom
            )

            # Right border
            right_cell = ws.cell(row=row, column=group_end)
            right_cell.border = Border(
                left=right_cell.border.left,
                right=thick,
                top=right_cell.border.top,
                bottom=right_cell.border.bottom
            )

def export_to_excel(dataframes, sheet_names, start_date, end_date):
    """
    Export dataframes to an Excel file with multiple sheets.
    """
    try:
        output = BytesIO()

        # Load the template workbook
        wb = load_workbook(TEMPLATE_PATH)

        for df, sheet_name in zip(dataframes, sheet_names):
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                st.error(f"Sheet '{sheet_name}' not found in the template.")
                continue

            # Clear existing data starting from row 4
            if ws.max_row >= 4:
                ws.delete_rows(4, ws.max_row)

            # Format header
            format_header(ws, sheet_name, start_date, end_date)

            # Write DataFrame to worksheet
            total_rows_written = write_dataframe_to_worksheet(ws, df)

            # Apply number formatting
            apply_number_formatting(ws, df)

            # Apply conditional formatting for totals
            apply_conditional_formatting(ws, df)

            # Apply performance formatting to all target columns
            apply_performance_formatting(ws, df, [
                "Services % to PY",
                "Retail % to PY",
                "Total % to PY", 
                "Services % to Budget",
                "Retail % to Budget",
                "Total % to Budget"
            ])

            # Merge BRAND cells with alignment adjustment
            merge_brand_cells(ws)

            # Apply gridlines with proper DataFrame passing
            apply_gridlines(ws, df)

        # Save the workbook to the output buffer
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        st.error(f"Error generating Excel file: {str(e)}")
        return None