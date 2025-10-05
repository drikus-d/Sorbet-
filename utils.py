'''
Utility functions for the Clicks Scorecard
'''
import streamlit as st
import pandas as pd
import config
from typing import Tuple, List
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.styles import Font

# Import Snowflake modules with error handling
try:
    from snowflake.snowpark import Session
    from snowflake.snowpark.context import get_active_session
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.backends import default_backend
    SNOWFLAKE_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Snowflake modules not available: {str(e)}")
    SNOWFLAKE_AVAILABLE = False
    # Create dummy classes to prevent NameError
    class Session:
        @staticmethod
        def builder():
            return None
    def get_active_session():
        raise Exception("Snowflake not available")

def get_snowflake_session():
    """
    Get the active Snowflake session, or create one if it doesn't exist.
    """
    if not SNOWFLAKE_AVAILABLE:
        st.error("Snowflake modules are not available. Please check your installation.")
        return None
        
    try:
        active_session = get_active_session()
        return active_session
    except Exception: # pylint: disable=broad-except
        try:
            private_key_der = get_private_key()
            # If no active session, create a new one
            connection_parameters = {
                "account": config.SNOWFLAKE_ACCOUNT,
                "user": config.SNOWFLAKE_USER,
                "private_key": private_key_der,
                "role": config.SNOWFLAKE_ROLE,
                "warehouse": config.SNOWFLAKE_WAREHOUSE,
                "database": config.SNOWFLAKE_DATABASE,
                "schema": config.SNOWFLAKE_SCHEMA,
            }

            session = Session.builder.configs(connection_parameters).create()
            return session
        except Exception as e:
            st.error(f"Failed to establish Snowflake session: {str(e)}")
            return None



def date_range_selection() -> Tuple[pd.Timestamp, pd.Timestamp]:
    """Allow the user to select the date range for the data with defaults set to yesterday and a week ago."""
    
    # Calculate yesterday's date
    yesterday = datetime.today().date() - timedelta(days=1)
    
    # Calculate the date one week before yesterday
    week_ago = yesterday - timedelta(days=7)
    
    # Create two columns in Streamlit for Start Date and End Date inputs
    col1, col2 = st.columns(2)
    
    with col1:
        # Start Date input with default value set to a week ago
        start_date = st.date_input(
            "Start Date",
            value=week_ago,
            min_value=None,  # No minimum date restriction
            max_value=None   # Removed max_value to allow any date selection
        )
    
    with col2:
        # End Date input with default value set to yesterday
        end_date = st.date_input(
            "End Date",
            value=yesterday,
            min_value=start_date,  # Ensure end_date is not before start_date
            max_value=None          # Removed max_value to allow any date selection
        )
    
    # Convert datetime.date objects to pandas Timestamps for consistency
    start_date = pd.Timestamp(start_date)
    end_date = pd.Timestamp(end_date)
    
    return start_date, end_date

def read_sql_query(file_path: str) -> str:
    """Read the SQL query from the provided file path."""
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            return file.read()
    except FileNotFoundError:
        st.error("SQL file not found. Please ensure the file path is correct.")
        return ""


@st.cache_data
def load_data(_session, query):
    """Load and process data, with caching to avoid re-execution on each run."""
    result = _session.sql(query).collect()
    df = pd.DataFrame(result)
    return df

def custom_formatter(x):
    '''
    Format the data for display.
    '''
    if isinstance(x, (int, float)):
        return f"{x:,.0f}"
    else:
        return x

def combine_dataframes_with_empty_rows(dataframes, subheaders, empty_rows=2):
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Data"

    current_row = 1

    for df, subheader in zip(dataframes, subheaders):
        # Write the subheader above each table
        header_cell = ws.cell(row=current_row, column=1, value=subheader)
        header_cell.font = Font(bold=True)
        current_row += 1

        # Ensure df is a plain DataFrame, not a Styler object
        if isinstance(df, pd.io.formats.style.Styler):
            df = df.data  # Get the underlying DataFrame

        # Write DataFrame headers and data to the Excel sheet
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=True), start=current_row
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Update the current row to add empty rows after each table
        current_row = r_idx + 1 + empty_rows

    # Save the workbook to a BytesIO stream
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)  # Reset stream position for reading
    return excel_stream


def generate_combined_dataframe(all_dataframes: List[pd.DataFrame], subheaders: List[str]) -> pd.DataFrame:
    """
    Generates a combined DataFrame listing all DataFrames, their subheaders, columns, and data types.
    
    Parameters:
        all_dataframes (List[pd.DataFrame]): List of DataFrames to include.
        subheaders (List[str]): List of subheader strings corresponding to each DataFrame.
    
    Returns:
        pd.DataFrame: Combined DataFrame with columns ['Subheader', 'Column Name', 'Data Type', 'Desired Format'].
    """
    combined_data = []
    for df, subheader in zip(all_dataframes, subheaders):
        for col in df.columns:
            data_type = str(df[col].dtype)
            combined_data.append({
                "Subheader": subheader,
                "Column Name": col,
                "Data Type": data_type,
                "Desired Format": ""  # To be filled by the user
            })
    combined_df = pd.DataFrame(combined_data)
    return combined_df

def format_column_name(column_name: str) -> str:
    """Converts a column name in the format SECTION-METRIC to 'Section Metric'."""
    parts = column_name.split('-')
    formatted_parts = [' '.join(word.capitalize() for word in part.split('_')) for part in parts]
    return ' - '.join(formatted_parts)


def rename_columns_with_spaces(df: pd.DataFrame) -> pd.DataFrame:
    """
    Renames columns in the dataframe by converting CAPS_SNAKE_CASE to Capitalized with spaces.
    """
    # Define a function to format each column name
    def format_column_name(column_name) -> str:
        # Convert column name to string and split by underscores
        return ' '.join(str(word).capitalize() for word in str(column_name).split('_'))
    
    # Apply the formatting function to all column names
    df.columns = [format_column_name(col) for col in df.columns]
    
    return df

def get_private_key():
    if not SNOWFLAKE_AVAILABLE:
        raise Exception("Snowflake modules are not available")
        
    # Local import to avoid undefined name if the top-level optional import failed
    from cryptography.hazmat.primitives import serialization as _serialization
    from cryptography.hazmat.backends import default_backend as _default_backend

    return _serialization.load_pem_private_key(
        config.SNOWFLAKE_PRIVATE_KEY.encode(),
        password=None,
        backend=_default_backend()
    ).private_bytes(
        encoding=_serialization.Encoding.DER,
        format=_serialization.PrivateFormat.PKCS8,
        encryption_algorithm=_serialization.NoEncryption()
    )